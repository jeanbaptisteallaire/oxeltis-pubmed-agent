"""
Oxeltis — PubMed Lead Finder
Identifies biotechs with active drug discovery programs via PubMed publications
"""

import streamlit as st
import requests
import anthropic
import pandas as pd
import time
import re
import json
from io import BytesIO
from xml.etree import ElementTree as ET
from datetime import datetime, timedelta
from urllib.parse import urlparse

# ── Import default values from config.py ─────────────────────────────────────
from config import (
    OXELTIS_CONTEXT      as DEFAULT_OXELTIS_CONTEXT,
    SCORING_RULES        as DEFAULT_SCORING_RULES,
    INDUSTRY_KEYWORDS    as DEFAULT_INDUSTRY_KEYWORDS,
    ACADEMIC_KEYWORDS    as DEFAULT_ACADEMIC_KEYWORDS,
    SKIP_DOMAINS         as DEFAULT_SKIP_DOMAINS,
)

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Oxeltis — PubMed Lead Finder",
    page_icon="🔬",
    layout="wide"
)

# ── Brand colors ──────────────────────────────────────────────────────────────
OXELTIS_PURPLE = "#3D2785"

# ── Custom CSS — purple sidebar, white text ───────────────────────────────────
st.markdown(f"""
<style>
/* Sidebar background */
[data-testid="stSidebar"] {{
    background-color: {OXELTIS_PURPLE};
}}
/* All text in sidebar → white */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] .stMarkdown {{
    color: white !important;
}}
/* Slider track */
[data-testid="stSidebar"] [data-testid="stSlider"] div {{
    color: white !important;
}}
/* Info/tip box in sidebar */
[data-testid="stSidebar"] .stAlert {{
    background-color: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.3);
    color: white !important;
}}
[data-testid="stSidebar"] .stAlert p {{
    color: white !important;
}}
/* Divider */
[data-testid="stSidebar"] hr {{
    border-color: rgba(255,255,255,0.3);
}}
/* Button in sidebar */
[data-testid="stSidebar"] .stButton button {{
    background-color: white;
    color: {OXELTIS_PURPLE};
    font-weight: bold;
    border: none;
}}
[data-testid="stSidebar"] .stButton button:hover {{
    background-color: #f0ecff;
    color: {OXELTIS_PURPLE};
}}
</style>
""", unsafe_allow_html=True)

# ── Session state — load defaults on first run ────────────────────────────────
if "oxeltis_context" not in st.session_state:
    st.session_state.oxeltis_context = DEFAULT_OXELTIS_CONTEXT
if "scoring_rules" not in st.session_state:
    st.session_state.scoring_rules   = DEFAULT_SCORING_RULES
if "industry_kw" not in st.session_state:
    st.session_state.industry_kw     = "\n".join(DEFAULT_INDUSTRY_KEYWORDS)
if "academic_kw" not in st.session_state:
    st.session_state.academic_kw     = "\n".join(DEFAULT_ACADEMIC_KEYWORDS)
if "skip_domains" not in st.session_state:
    st.session_state.skip_domains    = "\n".join(DEFAULT_SKIP_DOMAINS)


# ── Helpers ───────────────────────────────────────────────────────────────────

def build_system_prompt():
    return f"""You are an expert in biotech analysis for qualifying commercial prospects for Oxeltis.

{st.session_state.oxeltis_context}

Analyze the content of a biotech website and return ONLY a valid JSON object,
with no text before or after, no markdown, no backticks.

{st.session_state.scoring_rules}

Expected JSON format:
{{
  "modalite": "small_molecule|biologics|mixed|unknown",
  "stade": "hit_to_lead|lead_opt|preclinical|clinical|unknown",
  "score": <integer 0 to 5>,
  "accroche": "<one outreach sentence in English, personalized, mentioning the company and their program>"
}}"""


def get_industry_kw():
    return [k.strip() for k in st.session_state.industry_kw.splitlines() if k.strip()]

def get_academic_kw():
    return [k.strip() for k in st.session_state.academic_kw.splitlines() if k.strip()]

def get_skip_domains():
    return [d.strip() for d in st.session_state.skip_domains.splitlines() if d.strip()]


# ── PubMed ────────────────────────────────────────────────────────────────────

def pubmed_search(query, max_results=30, months_back=12):
    end_date   = datetime.now()
    start_date = end_date - timedelta(days=months_back * 30)
    date_filter = f"{start_date.strftime('%Y/%m/%d')}:{end_date.strftime('%Y/%m/%d')}[pdat]"
    full_query  = f"({query}) AND {date_filter}"
    try:
        r = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
            params={"db": "pubmed", "term": full_query, "retmax": max_results,
                    "retmode": "json", "sort": "relevance"},
            timeout=15
        )
        r.raise_for_status()
        return r.json().get("esearchresult", {}).get("idlist", [])
    except Exception as e:
        st.error(f"PubMed search error: {e}")
        return []


def pubmed_fetch(pmids):
    if not pmids:
        return []
    try:
        r = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
            params={"db": "pubmed", "id": ",".join(pmids), "retmode": "xml", "rettype": "abstract"},
            timeout=30
        )
        r.raise_for_status()
    except Exception as e:
        st.error(f"PubMed fetch error: {e}")
        return []

    papers = []
    try:
        root = ET.fromstring(r.text)
    except ET.ParseError:
        return []

    for article in root.findall(".//PubmedArticle"):
        title_el = article.find(".//ArticleTitle")
        title    = (title_el.text or "") if title_el is not None else ""
        abstract = " ".join([(el.text or "") for el in article.findall(".//AbstractText")])
        affiliations = [aff.text for aff in article.findall(".//Affiliation") if aff.text]
        pmid_el  = article.find(".//PMID")
        pmid     = pmid_el.text if pmid_el is not None else ""
        papers.append({"pmid": pmid, "title": title, "abstract": abstract[:400], "affiliations": affiliations})

    return papers


def extract_companies(affiliations):
    industry_kw = get_industry_kw()
    academic_kw = get_academic_kw()
    companies = []
    for aff in affiliations:
        aff_lower = aff.lower()
        if any(kw in aff_lower for kw in academic_kw):
            continue
        if any(kw in aff_lower for kw in industry_kw):
            name = re.sub(r'\s+', ' ', aff.split(",")[0].strip())
            if 3 < len(name) < 80:
                companies.append(name)
    return list(set(companies))


# ── Firecrawl ─────────────────────────────────────────────────────────────────

def firecrawl_search_url(company_name, api_key):
    skip = get_skip_domains()
    try:
        r = requests.post(
            "https://api.firecrawl.dev/v1/search",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"query": f"{company_name} official website biotech drug discovery", "limit": 5},
            timeout=15
        )
        if r.status_code == 200:
            for result in r.json().get("data", []):
                url = result.get("url", "")
                if not any(d in url for d in skip):
                    parsed = urlparse(url)
                    return f"{parsed.scheme}://{parsed.netloc}/"
    except Exception:
        pass
    return None


def firecrawl_scrape(url, api_key):
    try:
        r = requests.post(
            "https://api.firecrawl.dev/v1/scrape",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"url": url, "formats": ["markdown"], "onlyMainContent": True},
            timeout=20
        )
        if r.status_code == 200:
            content = r.json().get("data", {}).get("markdown", "")
            if content and len(content) > 100:
                return content[:2500]
    except Exception:
        pass
    return None


# ── Claude ────────────────────────────────────────────────────────────────────

def analyze_company(company_name, site_content, client, paper_title="", paper_abstract=""):
    system_prompt = build_system_prompt()
    user_msg = f"Company: {company_name}"
    if paper_title:
        user_msg += f"\nAssociated PubMed paper: {paper_title}"
    if paper_abstract:
        user_msg += f"\nPaper abstract (excerpt): {paper_abstract[:300]}"
    user_msg += f"\n\nWebsite content:\n{site_content}"

    for attempt in range(3):
        try:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                system=system_prompt,
                messages=[{"role": "user", "content": user_msg}]
            )
            raw    = re.sub(r"```(?:json)?", "", response.content[0].text.strip()).strip()
            result = json.loads(raw)
            assert "score" in result and "accroche" in result
            result["score"] = max(0, min(5, int(result["score"])))
            return result
        except Exception:
            if attempt < 2:
                time.sleep(2)

    return {"modalite": "unknown", "stade": "unknown", "score": 0,
            "accroche": "Analysis error — manual review required"}


# ── Excel export ──────────────────────────────────────────────────────────────

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PubMed Leads')
        ws = writer.sheets['PubMed Leads']
        from openpyxl.styles import PatternFill, Font
        green  = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        pink   = PatternFill("solid", fgColor="FFC7CE")
        gray   = PatternFill("solid", fgColor="EFEFEF")
        for cell in ws[1]:
            cell.font = Font(bold=True)
        score_col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Score"), None)
        if score_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                score = row[score_col - 1].value or 0
                fill  = green if score >= 4 else (yellow if score >= 2 else (pink if score == 1 else gray))
                for cell in row:
                    cell.fill = fill
        for col, width in {"A": 30, "B": 60, "C": 30, "D": 15, "E": 8, "F": 15, "G": 70, "H": 12}.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = ws.dimensions
    return output.getvalue()


def stars(score):
    return "☆☆☆☆☆" if score == 0 else "★" * score + "☆" * (5 - score)


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

# ── Header ────────────────────────────────────────────────────────────────────
col_logo, col_title = st.columns([1, 5])
with col_logo:
    try:
        st.image("logo.png", width=160)
    except Exception:
        pass
with col_title:
    st.title("PubMed Lead Finder")
    st.caption("Identifies biotechs with active drug discovery programs via recent scientific publications")

st.divider()

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_search, tab_config = st.tabs(["🔍 Search", "⚙️ Qualification Criteria"])


# ── TAB 2: Criteria ───────────────────────────────────────────────────────────
with tab_config:
    st.subheader("⚙️ Qualification Criteria — editable for this session")
    st.info("Changes apply to the next search only. Settings reset when you reload the page.")

    st.markdown("### 1. Oxeltis services description")
    st.caption("This text is sent to Claude so it knows what Oxeltis does and does not do.")
    new_context = st.text_area("Oxeltis context", value=st.session_state.oxeltis_context,
                                height=280, label_visibility="collapsed")

    st.markdown("### 2. Scoring rules (0 to 5)")
    st.caption("Defines what makes a hot prospect (5), warm (3), or out of scope (1). Adjust based on field feedback.")
    new_scoring = st.text_area("Scoring rules", value=st.session_state.scoring_rules,
                                height=340, label_visibility="collapsed")

    col_left, col_right = st.columns(2)
    with col_left:
        st.markdown("### 3. Keywords — INDUSTRY affiliations")
        st.caption("One keyword per line. If found in a PubMed author affiliation → company is kept for analysis.")
        new_industry = st.text_area("Industry keywords", value=st.session_state.industry_kw,
                                     height=260, label_visibility="collapsed")
    with col_right:
        st.markdown("### 4. Keywords — ACADEMIC affiliations (to exclude)")
        st.caption("One keyword per line. If found → affiliation is ignored (lab, university, hospital...).")
        new_academic = st.text_area("Academic keywords", value=st.session_state.academic_kw,
                                     height=260, label_visibility="collapsed")

    st.markdown("### 5. Domains to ignore")
    st.caption("One domain per line. These sites will never be retained as a company's official website.")
    new_skip = st.text_area("Domains to ignore", value=st.session_state.skip_domains,
                             height=200, label_visibility="collapsed")

    col_btn1, col_btn2, _ = st.columns([2, 2, 4])
    with col_btn1:
        if st.button("💾 Apply changes", type="primary", use_container_width=True):
            st.session_state.oxeltis_context = new_context
            st.session_state.scoring_rules   = new_scoring
            st.session_state.industry_kw     = new_industry
            st.session_state.academic_kw     = new_academic
            st.session_state.skip_domains    = new_skip
            st.success("✅ Criteria updated! The next search will use these settings.")
    with col_btn2:
        if st.button("↺ Reset to defaults", use_container_width=True):
            st.session_state.oxeltis_context = DEFAULT_OXELTIS_CONTEXT
            st.session_state.scoring_rules   = DEFAULT_SCORING_RULES
            st.session_state.industry_kw     = "\n".join(DEFAULT_INDUSTRY_KEYWORDS)
            st.session_state.academic_kw     = "\n".join(DEFAULT_ACADEMIC_KEYWORDS)
            st.session_state.skip_domains    = "\n".join(DEFAULT_SKIP_DOMAINS)
            st.rerun()


# ── TAB 1: Search ─────────────────────────────────────────────────────────────
with tab_search:

    with st.sidebar:
        try:
            st.image("logo.png", use_container_width=True)
        except Exception:
            st.markdown("### Oxeltis")

        st.markdown("---")
        st.markdown("### API Keys")
        anthropic_key = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-api03-...")
        firecrawl_key = st.text_input("Firecrawl API Key", type="password", placeholder="fc-...")

        st.markdown("---")
        st.markdown("### Search Parameters")
        search_query = st.text_input(
            "PubMed keywords",
            value="small molecule drug discovery hit-to-lead",
            help="Examples: 'antiviral nucleoside synthesis', 'PROTAC degrader oncology', 'medicinal chemistry lead optimization'"
        )
        max_papers  = st.slider("Papers to analyze", 10, 100, 30, step=10)
        months_back = st.slider("Time window (months)", 1, 24, 12)
        min_score   = st.slider("Minimum score to display", 0, 5, 3)

        st.markdown("---")
        st.info("💡 30 papers → ~10-20 companies → ~$0.10-0.20 in API costs.")

        run_button = st.button("🚀 Run search", type="primary", use_container_width=True)

    if run_button:
        if not anthropic_key or not firecrawl_key:
            st.error("⚠️ Please enter both API keys in the sidebar.")
            st.stop()

        client  = anthropic.Anthropic(api_key=anthropic_key)
        results = []

        # ── Step 1: PubMed ─────────────────────────────────────────────────────
        with st.status("🔍 Step 1/3 — PubMed search...", expanded=True) as status:
            st.write(f"Query: `{search_query}` — last {months_back} months")
            pmids = pubmed_search(search_query, max_papers, months_back)

            if not pmids:
                st.error("No papers found. Try different keywords.")
                st.stop()

            st.write(f"✅ {len(pmids)} papers found")
            papers = pubmed_fetch(pmids)
            st.write(f"✅ {len(papers)} papers retrieved with affiliations")

            all_companies = {}  # company_name → (paper_title, paper_abstract)
            for paper in papers:
                for c in extract_companies(paper["affiliations"]):
                    if c not in all_companies:
                        all_companies[c] = (paper["title"], paper["abstract"])

            n = len(all_companies)
            st.write(f"🏢 **{n} industry companies identified** in affiliations")
            status.update(label=f"✅ PubMed done — {n} companies found", state="complete")

        if not all_companies:
            st.warning("No companies detected. Affiliations may be entirely academic. Adjust keywords in the ⚙️ Criteria tab.")
            st.stop()

        # ── Step 2 & 3: Firecrawl + Claude ────────────────────────────────────
        st.info(f"🔄 Analyzing {n} companies via Firecrawl + Claude Haiku...")
        progress_bar   = st.progress(0)
        status_text    = st.empty()
        companies_list = list(all_companies.items())

        for i, (company, (paper_title, paper_abstract)) in enumerate(companies_list):
            progress_bar.progress((i + 1) / len(companies_list))
            status_text.text(f"[{i+1}/{len(companies_list)}] {company}...")

            url = firecrawl_search_url(company, firecrawl_key)
            time.sleep(0.5)

            if not url:
                results.append({
                    "Company": company,
                    "PubMed Paper": paper_title[:80] + ("..." if len(paper_title) > 80 else ""),
                    "Website": "", "Modality": "unknown", "Score": 0, "Stage": "unknown",
                    "Outreach": "URL not found — manual prospecting required", "★": "☆☆☆☆☆"
                })
                continue

            content = firecrawl_scrape(url, firecrawl_key)
            time.sleep(0.5)

            if not content:
                results.append({
                    "Company": company,
                    "PubMed Paper": paper_title[:80] + ("..." if len(paper_title) > 80 else ""),
                    "Website": url, "Modality": "unknown", "Score": 0, "Stage": "unknown",
                    "Outreach": "Website inaccessible — manual review required", "★": "☆☆☆☆☆"
                })
                continue

            r = analyze_company(company, content, client, paper_title, paper_abstract)
            time.sleep(0.3)

            results.append({
                "Company": company,
                "PubMed Paper": paper_title[:80] + ("..." if len(paper_title) > 80 else ""),
                "Website": url,
                "Modality": r["modalite"],
                "Score": r["score"],
                "Stage": r["stade"],
                "Outreach": r["accroche"],
                "★": stars(r["score"])
            })

        progress_bar.progress(1.0)
        status_text.text("✅ Analysis complete!")

        # ── Results ────────────────────────────────────────────────────────────
        df     = pd.DataFrame(results)
        df_all = df.sort_values("Score", ascending=False)
        df_flt = df_all[df_all["Score"] >= min_score]

        st.divider()
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Companies analyzed", len(results))
        c2.metric(f"Score ≥ {min_score} ★", len(df_flt))
        c3.metric("Score 5 ★", len(df_all[df_all["Score"] == 5]))
        c4.metric("Score 4 ★", len(df_all[df_all["Score"] == 4]))

        st.subheader(f"🎯 Qualified leads (Score ≥ {min_score})")

        if df_flt.empty:
            st.warning("No leads at this minimum score. Lower the filter in the sidebar.")
        else:
            def highlight_score(row):
                s = row["Score"]
                color = "#C6EFCE" if s >= 4 else ("#FFEB9C" if s >= 2 else ("#FFC7CE" if s == 1 else "#EFEFEF"))
                return [f"background-color: {color}"] * len(row)

            st.dataframe(df_flt.style.apply(highlight_score, axis=1), use_container_width=True, height=400)

            st.download_button(
                label="📥 Download Excel (qualified leads)",
                data=to_excel(df_flt),
                file_name=f"oxeltis_pubmed_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            top = df_flt[df_flt["Score"] >= 4]
            if not top.empty:
                st.subheader(f"🏆 Top Prospects ({len(top)} companies ≥ 4★)")
                for _, row in top.iterrows():
                    with st.expander(f"{row['★']} **{row['Company']}** — {row['Modality']} / {row['Stage']}"):
                        if row["Website"]:
                            st.write(f"**Website:** {row['Website']}")
                        st.write(f"**Paper:** {row['PubMed Paper']}")
                        st.write(f"**Outreach:** {row['Outreach']}")

        with st.expander("View all analyzed companies"):
            st.dataframe(df_all, use_container_width=True)
            st.download_button(
                label="📥 Download Excel (all companies)",
                data=to_excel(df_all),
                file_name=f"oxeltis_pubmed_all_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
